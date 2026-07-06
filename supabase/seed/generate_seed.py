#!/usr/bin/env python3
"""Generate seed.sql from the household spreadsheet.

Reads tmp/Financial Report 2026.xlsx and emits INSERT statements for the Kita
schema. Money is stored as integer cents. Auth user references are left as
:CH_UID / :JC_UID placeholders, substituted at apply time once the Supabase
auth users exist.
"""
import openpyxl, datetime, re, uuid

HH = '00000000-0000-0000-0000-0000000000aa'  # fixed household id
# Fixed namespace for deterministic vendor/location UUIDs (stable across re-runs).
SEED_NS = uuid.UUID('00000000-0000-0000-0000-0000000000aa')
wb = openpyxl.load_workbook('tmp/Financial Report 2026.xlsx', data_only=True)
out = []
# phase4 = incremental additions (vehicle transactions + ledger_entries) for
# households that already applied the Phase 1 seed. Also folded into `out`.
phase4 = []


def c(v):  # to cents; non-numeric -> 0
    if v is None:
        return 0
    try:
        return round(float(v) * 100)
    except (TypeError, ValueError):
        return 0


def s(v):  # sql string literal or NULL
    return 'NULL' if v is None else "'" + str(v).replace("'", "''") + "'"


def d(v):  # date literal or NULL
    if isinstance(v, (datetime.datetime, datetime.date)):
        return "'" + v.strftime('%Y-%m-%d') + "'"
    return 'NULL'


def is_num(v):  # numeric payment/amount cell, excluding bool and text like 'CLOSED'
    return isinstance(v, (int, float)) and not isinstance(v, bool)


_DATE_RE = re.compile(r'^\s*(\d{1,2})/(\d{1,2})/(\d{4})')


def parse_flex_date(v):
    """Parse a datetime cell OR a 'dd/mm/yyyy[...]' string (allowing trailing
    text such as '23/4/2024 (battery)'). Returns (iso_date_str_or_None, extra_text)."""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime('%Y-%m-%d'), ''
    if isinstance(v, str):
        m = _DATE_RE.match(v)
        if m:
            day, month, year = (int(x) for x in m.groups())
            try:
                iso = datetime.date(year, month, day).strftime('%Y-%m-%d')
            except ValueError:
                return None, ''
            return iso, v[m.end():].strip()
    return None, ''


out.append(f"insert into households(id,name) values ('{HH}','Chong Family');")
out.append(f"insert into household_members(household_id,user_id,role,member_code) values"
           f" ('{HH}',:CH_UID,'owner','CH'),('{HH}',:JC_UID,'member','JC');")

# Joint Fund config: CH expected 2270/mo, JC 2470/mo; JC carry-forward 1338.53
out.append(f"insert into joint_fund_config(household_id,member_code,expected_monthly_cents,carry_forward_prev_year_cents) values"
           f" ('{HH}','CH',227000,0),('{HH}','JC',247000,133853);")

# Joint Fund contributions — two column groups in 'Joint Fund' sheet
jf = wb['Joint Fund']
for col_date, col_amt, col_status, member in [(1, 2, 3, 'CH'), (7, 8, 9, 'JC')]:
    for r in range(3, 15):
        dt = jf.cell(r, col_date).value
        amt = jf.cell(r, col_amt).value
        st = jf.cell(r, col_status).value
        if dt is None or amt is None:
            continue
        status = 'paid' if st is True else 'pending'
        out.append(f"insert into joint_fund_contributions(household_id,member_code,period,amount_cents,status) "
                   f"values ('{HH}','{member}',{d(dt)},{c(amt)},'{status}');")

# Budget categories — 'Money breakdown' rows 2..8 (Category, JC, CH, Total, Remark)
mb = wb['Money breakdown']
order = 0
for r in range(2, 9):
    name = mb.cell(r, 1).value
    if not name:
        continue
    order += 1
    out.append(f"insert into budget_categories(household_id,name_en,jc_cents,ch_cents,total_cents,remark,sort_order) "
               f"values ('{HH}',{s(name)},{c(mb.cell(r, 2).value)},{c(mb.cell(r, 3).value)},"
               f"{c(mb.cell(r, 4).value)},{s(mb.cell(r, 5).value)},{order});")

# Monthly commitments — 'Money breakdown' second table rows 2..9 (cols 7,8,9)
for r in range(2, 10):
    name = mb.cell(r, 7).value
    amt = mb.cell(r, 8).value
    if not name or amt is None:
        continue
    out.append(f"insert into monthly_commitments(household_id,name_en,amount_cents,remark) "
               f"values ('{HH}',{s(name)},{c(amt)},{s(mb.cell(r, 9).value)});")

# Expenses — 'Expenses' sheet (Date,Vendor,Location,Details,Amount)
# expenses.date is NOT NULL; the sheet leaves some rows undated (they belong to
# the same period as the dated row above). Forward-fill the last seen date so no
# row is dropped and each stays in the correct month.
#
# Vendor/location are now FK columns (0005_expense_catalog.sql dropped the old
# text columns), so: collect distinct non-blank names, emit vendors/locations
# rows with stable (deterministic) ids, then link each expense by id. '-' is
# treated as blank/no-location, matching the migration's backfill logic.
# Categories are intentionally NOT seeded (the Excel has none).
ex = wb['Expenses']


def blank_or_dash(v):
    return v is None or str(v).strip() in ('', '-')


def stable_id(kind, name):
    # Deterministic uuid5 from a fixed namespace so re-running the generator
    # yields identical ids (stable FK links, idempotent regeneration).
    return str(uuid.uuid5(SEED_NS, f'{HH}:{kind}:{str(name).strip().lower()}'))


# Pass 1: read all expense rows once, collecting distinct vendor/location names.
vendor_ids, vendor_names = {}, {}      # lower(name) -> id / display name
location_ids, location_names = {}, {}
expense_rows = []
last_date = None
for r in range(2, ex.max_row + 1):
    amt = ex.cell(r, 5).value
    if amt is None:
        continue
    dt = ex.cell(r, 1).value
    if isinstance(dt, (datetime.datetime, datetime.date)):
        last_date = dt
    vendor = ex.cell(r, 2).value
    location = ex.cell(r, 3).value
    details = ex.cell(r, 4).value
    expense_rows.append((last_date, vendor, location, details, amt))

    if not blank_or_dash(vendor):
        key = str(vendor).strip().lower()
        vendor_names.setdefault(key, str(vendor).strip())
        vendor_ids.setdefault(key, stable_id('vendor', vendor))
    if not blank_or_dash(location):
        key = str(location).strip().lower()
        location_names.setdefault(key, str(location).strip())
        location_ids.setdefault(key, stable_id('location', location))

for key in sorted(vendor_names):
    out.append(f"insert into vendors (id,household_id,name) values ('{vendor_ids[key]}','{HH}',{s(vendor_names[key])});")
for key in sorted(location_names):
    out.append(f"insert into locations (id,household_id,name) values ('{location_ids[key]}','{HH}',{s(location_names[key])});")

# Pass 2: emit expenses linked by vendor_id/location_id (blank/'-' -> NULL).
for last_date, vendor, location, details, amt in expense_rows:
    vendor_id = 'NULL' if blank_or_dash(vendor) else f"'{vendor_ids[str(vendor).strip().lower()]}'"
    location_id = 'NULL' if blank_or_dash(location) else f"'{location_ids[str(location).strip().lower()]}'"
    out.append(f"insert into expenses(household_id,date,vendor_id,location_id,details,amount_cents,paid_by) "
               f"values ('{HH}',{d(last_date)},{vendor_id},{location_id},"
               f"{s(details)},{c(amt)},NULL);")

# Assets: TreeO (property), Myvi/Alza (vehicle), AIA-CH/AIA-JC (investment)
A_TREEO = '00000000-0000-0000-0000-0000000000b1'
A_MYVI = '00000000-0000-0000-0000-0000000000b2'
A_ALZA = '00000000-0000-0000-0000-0000000000b3'
A_AIACH = '00000000-0000-0000-0000-0000000000b4'
A_AIAJC = '00000000-0000-0000-0000-0000000000b5'


def asset(idlit, typ, name, owner, opening):
    owner_sql = 'NULL' if owner is None else f"'{owner}'"
    open_sql = 'NULL' if opening is None else str(opening)
    out.append(f"insert into assets(id,household_id,type,name,owner_member_code,opening_balance_cents) "
               f"values ('{idlit}','{HH}','{typ}',{s(name)},{owner_sql},{open_sql});")


# TreeO opening balance = carry-forward from 2025 (TreeO sheet R2 = 46013.81)
asset(A_TREEO, 'property', 'TreeO', None, 4601381)
asset(A_MYVI, 'vehicle', 'Myvi PQC 9059', None, None)
asset(A_ALZA, 'vehicle', 'Alza PNM 9059', None, None)
asset(A_AIACH, 'investment', 'AIA — CH', 'CH', None)
asset(A_AIAJC, 'investment', 'AIA — JC', 'JC', None)

# TreeO transactions — 'TreeO' sheet rows 3..18 (Date,Details,Amount,Transferred)
to = wb['TreeO']
for r in range(3, 19):
    det = to.cell(r, 2).value
    amt = to.cell(r, 3).value
    if amt is None:
        continue
    transferred = to.cell(r, 4).value is True
    direction = 'in' if 'Commitment' in str(det) else 'out'
    txn_type = 'monthly_commitment' if direction == 'in' else 'bill'
    out.append(f"insert into asset_transactions(asset_id,household_id,date,description,amount_cents,direction,txn_type,settled) "
               f"values ('{A_TREEO}','{HH}',{d(to.cell(r, 1).value)},{s(det)},{c(amt)},'{direction}','{txn_type}',{str(transferred).lower()});")

# AIA schedules — 'AIA Investment' rows 3..12 (seq col1, CH date/amt col2/3, JC date/amt col5/6)
aia = wb['AIA Investment']
for r in range(3, 13):
    seq = aia.cell(r, 1).value
    if seq is None:
        continue
    seq = int(seq)
    for dcol, acol, aid in [(2, 3, A_AIACH), (5, 6, A_AIAJC)]:
        amt = aia.cell(r, acol).value
        if amt is None:
            continue
        out.append(f"insert into asset_transactions(asset_id,household_id,date,description,amount_cents,direction,txn_type,settled,seq) "
                   f"values ('{aid}','{HH}',{d(aia.cell(r, dcol).value)},'AIA payment',{c(amt)},'out','scheduled_payment',true,{seq});")

# Vehicle transactions — 'Car' sheet. Two blocks (merged title rows):
#   Myvi PQC 9059 = cols A1:H1  -> Bank/Loan (A-B), Road Tax + Insurance (D-E), Maintenance (G-H)
#   Alza PNM 9059 = cols J1:Q1  -> Loan Payback (J-K), Road Tax + Insurance (M-N), Maintenance (P-Q)
# Data rows start at row 4 (rows 1-3 are title/sub-header/column-header rows).
car = wb['Car']


def vehicle_txns(asset_id, blocks):
    for txn_type, label, dcol, acol in blocks:
        for r in range(4, 20):
            amt = car.cell(r, acol).value
            if not is_num(amt):
                continue  # skip 'CLOSED', blanks, headers
            raw_date = car.cell(r, dcol).value
            iso, extra = parse_flex_date(raw_date)
            if iso is None:
                continue  # asset_transactions.date is NOT NULL -> skip unparseable rows
            desc = f"{label} {extra}".strip()
            stmt = (f"insert into asset_transactions(asset_id,household_id,date,description,amount_cents,direction,txn_type,settled) "
                    f"values ('{asset_id}','{HH}','{iso}',{s(desc)},{c(amt)},'out','{txn_type}',true);")
            out.append(stmt)
            phase4.append(stmt)


vehicle_txns(A_MYVI, [
    ('loan', 'Bank/Loan', 1, 2),
    ('road_tax_insurance', 'Road Tax + Insurance', 4, 5),
    ('maintenance', 'Maintenance', 7, 8),
])
vehicle_txns(A_ALZA, [
    ('loan_payback', 'Loan Payback', 10, 11),
    ('road_tax_insurance', 'Road Tax + Insurance', 13, 14),
    ('maintenance', 'Maintenance', 16, 17),
])

# Ledger entries — CH/JC 'Personal' sheets. Each monthly block starts with a
# date in col 1; income = col2 desc (not 'Total') + col3 numeric amount
# (remark col4); expense = col6 desc (not 'Total'/'Balance') + col7 numeric
# amount (remark col8). No :CH_UID/:JC_UID placeholders — keyed on member_code.
def personal_ledger(sheet, member_code):
    current_period = None
    for r in range(1, sheet.max_row + 1):
        dt = sheet.cell(r, 1).value
        if isinstance(dt, (datetime.datetime, datetime.date)):
            current_period = dt
        if current_period is None:
            continue

        idesc = sheet.cell(r, 2).value
        iamt = sheet.cell(r, 3).value
        if idesc and str(idesc).strip() != 'Total' and is_num(iamt):
            stmt = (f"insert into ledger_entries(household_id,owner_member_code,period,entry_type,description,amount_cents,remark) "
                    f"values ('{HH}','{member_code}',{d(current_period)},'income',{s(idesc)},{c(iamt)},{s(sheet.cell(r, 4).value)});")
            out.append(stmt)
            phase4.append(stmt)

        edesc = sheet.cell(r, 6).value
        eamt = sheet.cell(r, 7).value
        if edesc and str(edesc).strip() not in ('Total', 'Balance') and is_num(eamt):
            stmt = (f"insert into ledger_entries(household_id,owner_member_code,period,entry_type,description,amount_cents,remark) "
                    f"values ('{HH}','{member_code}',{d(current_period)},'expense',{s(edesc)},{c(eamt)},{s(sheet.cell(r, 8).value)});")
            out.append(stmt)
            phase4.append(stmt)


personal_ledger(wb['CH (Personal)'], 'CH')
personal_ledger(wb['JC (Personal)'], 'JC')

with open('supabase/seed/seed-phase4.sql', 'w') as f:
    f.write('\n'.join(phase4) + '\n')

print('\n'.join(out))
